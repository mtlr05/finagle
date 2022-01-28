import pandas as pd
import xlrd
from string import ascii_letters 
import datetime
import numpy as np
from scipy import interpolate
import logging
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os

class company:
    '''
    This is a class to model the financials of a publicly traded company.
    It has a number of methods which should be run in order:
    __init__(), will load the data into the appropriate attributes and
    fcf_from_X(), there are several of these methods. They are all meant to be used for calculating fcf
    fcf_to_X(), these are methods which are used to model how the FCF is to be used: paying down debt, buying back shares
    value(), this method is use for calculating a DCF from the cashflows, This shold be run only once other methods described have been run
    display_fin(), this method is used to process the financials. It should be used once all modelling is completed
    '''
    def __init__(self, financials = None, ticker = None, re = None, rd = None, t = None, shares = 1, price = 0, gt = 0, fcfe = None, fcff = None, fcf = None, roict = 0.15, year = 6, dividend = 0):
        
        '''
        self = object instance of the current 
        financials = python dictionary of financial input data. each value can be a single value or a list... 
               First value in the 0 index location shoulds be the value which corresponds to the trailing twelve months (ttm)... 
               actual keys required by the object will depend on which subsequent methods are used. See methods doc string for specific guidance
        ticker = string of the ticker symbol of the company
        re = cost of equity
        rd = cost of debt
        t = marginal tax rate
        shares = shares outstanding. Should include all classes if multiple classes exhist
        gt = terminal growth, required for model closure
        year = final forecast year. Year before the terminal year.
        fcfe = free cash flow to equity
        fcff = free cash flow to firm
        roict = return on invested capital in the terminal year, 15% default, mostly creates impact via taxes (by calculating depreciation) since gt and capex are explicitly specified
        dividend = current dividend policy
        '''
        
        #setup logging
        self.logfile = ticker + '.log'
        logging.basicConfig(filemode='w', level=logging.INFO, format='%(asctime)s %(levelname)s:%(message)s')
        logging.FileHandler(filename = self.logfile,mode='w')
        logging.info(ticker)
        
        #read in various attributes
        self.ticker = ticker
        self.gt = gt
        self.re = re
        self.rd = rd
        self.t = t
        self.roict = roict
        self.shares = shares
        self.price = price
        self.year = year
        self.years = list(range(year+1))
        self.now = datetime.date.today()
        self.buybacks = False
        
        if isinstance(dividend,list):
            self.dividend = dividend
        elif isinstance(dividend,float):
            self.dividend = [dividend]
        elif isinstance(dividend,int):
            self.dividend = [dividend]

        #create financial dataframe
        if financials == None:
            pass
        else:
            self.load_financials(financials = financials)
            # financials['date'] = [datetime.datetime.strptime(financials['date'], '%Y-%m-%d')+datetime.timedelta(days=365*i) for i in range(self.year+1)]
            # self.fin = pd.DataFrame({key:pd.Series(value) for key,value in financials.items()})
            # self.fin.set_index('date',inplace=True)
            # self.fin['shares'] = shares
            # self.fin['price'] = price
            # self.fin['MnA'] = 0
            # self.fin['buybacks'] = 0
            # self.fin['cashBS'] = 0
            # self.fin['cashBS'].iloc[0] = self.fin['cash'].iloc[0] 
            # self.cash0 = self.fin['cash'].iloc[0] 
            
            # #check financial data completeness
            # self.__datacheck()
            # logging.info('input data used for the forecast is:')
            # logging.info(financials)
                
            #initialize the FCF columns, if you want to do a DCF directly without calculating from financials, this requires a financial dict with a date for the ttm year
            self.fin['fcfe'] = fcfe
            self.fin['fcff'] = fcff
            self.fin['fcf'] = fcf
            self.fin['dividend'] = (self.fin['fcfe']-self.fin['buybacks'])/self.fin['shares']



    def __stream(self,sf,st):
        '''
        create a periodic stream of values based of yearly forecasts and terminal values
        sf = forecasted stream values
        st = stream value in terminal year
        '''
        
        if np.isnan(st) == True:  
            logging.error('Terminal value is a NaN, cannot create forecast')
                 
        if isinstance(sf,list):
            length = np.count_nonzero(~np.isnan(sf))
        elif isinstance(sf,float):
            length = 1
            sf = [sf]
        else:
            length = np.count_nonzero(~np.isnan(sf))
            sf = list(sf.iloc[0:length])
        
        x=list(range(length))
        x.append(self.year)
        s=sf+[st]
        f=interpolate.interp1d(x,s)
        values = f(self.years)
        return values
    
    def __datacheck(self):
        '''check for all required columns,
        check sufficient number of years
        create flags for which fcf_from_x functions may be used'''
        
        self.data_for_earnings = False
        self.data_for_ebitda = False
        
        if 'ebitda' in self.fin.columns:
            if self.year < len(self.fin.ebitda)-1:
                self.year = len(self.fin.ebitda)-1
                self.years = list(range(self.year+1))
                logging.warning("Warning: length of EBITDA forecast appear larger then the 'year' parameter used at initialization")
        
            from_ebitda_columns = ['price', 'tax', 'interest', 'capex', 'noa', 'nol', 'ebitda', 'shares', 'dwc', 'debt', 'cash', 'da','MnA','buybacks','cashBS']
            from_ebitda_forecasts = ['ebitda','capex','dwc','debt']
        
            check_columns = set(from_ebitda_columns) == set(list(self.fin.columns.values))
            check_forecasts = np.count_nonzero(np.isnan(self.fin[from_ebitda_forecasts])) == 0
            self.data_for_ebitda = check_columns*check_forecasts
        
        if 'e' in self.fin.columns:
            self.data_for_earnings = True
                    
        if self.data_for_ebitda == True:
            logging.info('datacheck complete: financial dataset appears complete for fcf_from_ebitda')
        
        if self.data_for_earnings == True:
            logging.info('datacheck complete: financial dataset appears complete for fcf_from_earnings')
        
    def __pv(self,cfs,g,r,cft = None):
        '''
        calculate the present value of future cash flows. To be even more clear, any flows from the current year are not included in the present value calculation - only future years
        the last cf in cfs is the year before the terminal year
        cfs = array of cash flows
        cft = termal cash flow
        '''
        if isinstance(r,float):
            r = [r for cf in cfs]
        else:
            r = list(r)
        
        if cft == None:
            value = [cfs.iloc[-1]*(1+g)/(r[-1]-g)]
        else:
            value = [cft/(r[-1]-g)]
        r = r[::-1]
        for i,cf in enumerate(cfs[:0:-1],1):
            value.insert(0,(cf+value[0])/(1+r[i]))
        return value
    
    def __wacc(self):
        '''return weighted average cost of capital'''
        self.fin['EV'] = self.fin.debt + self.fin.equity  
        self.fin['wacc'] = (self.fin.debt*self.rd*(1-self.t)+self.fin.equity*self.re)/self.fin.EV
        return self.fin['wacc']
        
    def forecast_ebitda(self,ebitda_ttm,gf,financials = None):
        '''creates an ebitda forecast and populates the financials'''
        g = self.__stream(gf,self.gt)
        ebitda = [ebitda_ttm]
        for i in range(self.year):
            ebitda.append(ebitda[i]*(1+g[i]))
            
        if financials != None:
            financials['ebitda'] = ebitda #since this is a dict it will populate the elements similar to a pointer
        
        return ebitda
        
    def forecast_capex(self,capex_f,financials):
        '''creates an capex forecast and populates the financials'''
        if isinstance(capex_f,list):
            length = np.count_nonzero(~np.isnan(capex_f))
            capex = capex_f
        elif isinstance(capex_f,float):
            length = 1
            capex = [capex_f]
        elif isinstance(capex_f,int):
            length = 1
            capex = [capex_f]
        else:
            length = np.count_nonzero(~np.isnan(capex_f))
            capex = list(capex_f.iloc[0:length])

        for i in range(length,self.year+1): #start scaling with ebitda where the forecast ends
            capex.append(capex[length-1]/financials['ebitda'][length-1]*financials['ebitda'][i])
            
        if financials != None:
            financials['capex'] = capex #since this is a dict it will populate the elements similar to a pointer
        
        return capex    
        
    def load_financials(self, financials):
        financials['date'] = [datetime.datetime.strptime(financials['date'], '%Y-%m-%d')+datetime.timedelta(days=365*i) for i in range(self.year+1)]
        self.fin = pd.DataFrame({key:pd.Series(value) for key,value in financials.items()})
        self.fin.set_index('date',inplace=True)
        self.fin['shares'] = self.shares
        self.fin['price'] = self.price
        self.fin['MnA'] = 0
        self.fin['buybacks'] = 0
        self.fin['cashBS'] = 0
        
        try:
            self.fin['cashBS'].iloc[0] = self.fin['cash'].iloc[0] 
            self.cash0 = self.fin['cash'].iloc[0]
        except:
            logging.info('no cash key')
        
        self.__datacheck()
        logging.info('input data used for the forecast is:')
        logging.info(financials)
        logging.info('load_financials() method complete')
        
    def fcf_from_earnings(self, payout = 1 ,gf = 0, ROE = 1):
        '''
        payout = list of yearly payout forecasts, or most recent years
        ROE = perpetual return on equity
        gf = list of yearly growth forecast's 
        '''
        if self.data_for_earnings == False:
            logging.error('financial dataset cannot be used for calculating FCF from earnings')
        
        g = self.__stream(gf,self.gt) 

        for i in range(self.year):
            self.fin['e'].iloc[i+1]=self.fin['e'].iloc[i]*(1+g[i])
    
        payout_t = 1 - self.gt/ROE
        payouts = self.__stream(payout,payout_t)
    
        self.fin['fcfe'] = self.fin['e']*payouts
        logging.info('fcf_from_earnings() method complete')
        
    def fcf_from_ebitda(self):
        if self.data_for_ebitda == False:
            logging.error('financial dataset cannot be used for calculating FCF from EBITDA')
        
        interest0 = self.fin['interest'].iloc[0]
        self.fin['interest'] = self.rd*self.fin.debt.shift(1)
        self.fin['interest'].iloc[0] = interest0
        
        #really complicated way to calculate the terminal depreciation for situtions... 
        #...where there is terminal growth. This will enforce that Capex>=Depreciation....
        #...so that assets continue to increase as the company grows its bottom line  
        C=self.gt/self.roict*(1-self.t)
        dat=(self.fin['capex'].iloc[-1]-C*self.fin['ebitda'].iloc[-1])/(1-C)        
        if dat<0: 
            logging.error('negative depreciation in terminal year, check roic and growth assumptions')
        
        self.fin['da'] = self.__stream(self.fin['da'],dat) 
        self.fin['income_pretax'] = self.fin.ebitda - self.fin.da - self.fin.interest
        self.fin['dDebt'] = self.fin['debt']-self.fin['debt'].shift(1)
        self.fin['dDebt'].iloc[0] = 0 #todo: calculate from interest0 and Debt0
        
        
        self.fin['tax_cash'] = np.nan
        self.fin['tax_cash'].iloc[0] = self.fin['tax'].iloc[0]
        self.fin['income_taxable']= np.nan
        self.fin['income_taxable'].iloc[0] = max(self.fin['income_pretax'].iloc[0]*(1-self.fin['nol'].iloc[0]>0),0)
        for i in range(1,self.year+1):
            self.fin['nol'].iloc[i] = max(self.fin['nol'].iloc[i-1] - self.fin['income_pretax'].iloc[i],0)
            self.fin['income_taxable'].iloc[i] = max(0,self.fin['income_pretax'].iloc[i] - self.fin['nol'].iloc[i-1])
            self.fin['tax_cash'].iloc[i] = self.fin['tax_cash'].iloc[i-1]+self.t*(self.fin['income_taxable'].iloc[i]-self.fin['income_taxable'].iloc[i-1])
            self.fin['tax'].iloc[i] = self.fin['tax'].iloc[i-1]+self.t*(self.fin['income_pretax'].iloc[i]-self.fin['income_pretax'].iloc[i-1])

        self.fin['fcf'] = self.fin.ebitda - self.fin.tax_cash - self.fin.capex - self.fin.dwc - self.fin.interest
        self.fin['fcfe'] = self.fin.ebitda - self.fin.tax_cash - self.fin.capex - self.fin.dwc + self.fin.dDebt - self.fin.interest - self.fin.MnA
        self.fin['fcff'] = self.fin.ebitda - self.fin.tax_cash - self.fin.capex - self.fin.dwc - self.fin.interest*self.t - self.fin.MnA
        
        self.fin['dividend_policy'] = 0
        n_div = len(self.dividend)
        for i in range(self.year+1):
            if (i < n_div):
                self.fin['dividend_policy'].iloc[i] = self.dividend[i]*self.fin['shares'].iloc[i]
            else:
                self.fin['dividend_policy'].iloc[i] = max(self.dividend[n_div-1]*self.fin['shares'].iloc[n_div-1]/self.fin['fcf'].iloc[n_div-1]*self.fin['fcf'].iloc[i],self.fin['dividend_policy'].iloc[i-1])
            
        # for i in range(1,self.year+1):
            # self.fin['dividend_policy'].iloc[i] = max(self.dividend*self.shares/self.fin['fcf'].iloc[0]*self.fin['fcf'].iloc[i],self.fin['dividend_policy'].iloc[i-1])

        self.fin['dividend'] = (self.fin['fcfe']-self.fin['buybacks'])/self.fin['shares']
        self.fin['cash'].iloc[1:]  = self.fin['fcfe'].iloc[1:]
        self.fin['cash'] = self.fin['cash'].cumsum()
        self.fin['noa'].iloc[1:]  = self.fin['noa'].iloc[0]
        logging.info('fcf_from_ebitda() method complete')
        
    def fcf_to_debt(self,leverage = 3, year_d = 1):
        '''
        Adjust debt levels to desired target. Decrease (or increase) FCFE to reduce (or increase) debt towards target leverage
        prerequisite: Must first have fcf defined
        leverage = desired Debt/EBITDA, default value is 3
        '''
        #increase debt if fcf is negative and cash is 0
        if self.data_for_ebitda == False: 
            logging.error('financial dataset cannot be used to optimize leverage')
        
        if self.fin['fcf'].empty:
            logging.error('first calculate fcf')
        
        
        self.fin['debt_Target'] = [leverage*self.fin['ebitda'].iloc[i] if self.fin['ebitda'].iloc[i]>0 else 0 for i in range(self.year+1)] 

        #run the loop 3 times just to converge on the interest and FCF
        for i in range(3):
            for i in range(year_d-1, self.year):
                if (self.fin['debt'].iloc[i]-self.fin['debt_Target'].iloc[i+1]<0): #underlevered
                    dDebt = -1*(self.fin['debt'].iloc[i]-self.fin['debt_Target'].iloc[i+1]) 
                else: #overlevered
                    if i==0:
                        dDebt = -1*min(self.fin['debt'].iloc[i]-self.fin['debt_Target'].iloc[i+1],self.fin['fcf'].iloc[i+1]+self.fin['cash'].iloc[i]-self.fin['MnA'].iloc[i+1]-self.fin['dividend_policy'].iloc[i+1])
                    else:
                        dDebt = -1*min(self.fin['debt'].iloc[i]-self.fin['debt_Target'].iloc[i+1],self.fin['fcf'].iloc[i+1]-self.fin['MnA'].iloc[i+1]-self.fin['dividend_policy'].iloc[i+1])    
                self.fin['debt'].iloc[i+1] = self.fin['debt'].iloc[i]+dDebt
            self.fcf_from_ebitda()
        logging.info('fcf_to_debt() method complete')
    
    def fcf_to_bs(self):
        '''
        '''
        self.fin['cashBS'].iloc[0] = self.fin['cash'].iloc[0] 
        for i in range(self.year):
                self.fin['cashBS'].iloc[i+1] = self.fin['fcfe'].iloc[i+1] + self.fin['cashBS'].iloc[i] - self.fin['dividend_policy'].iloc[i+1]

        self.fin['dividend'] = self.fin['dividend_policy']/self.fin['shares']
        self.fin['dividend'].iloc[-1] = self.fin['dividend'].iloc[-1] + self.fin['cashBS'].iloc[-1]/self.fin['shares'].iloc[-1]
        self.cash0 = 0 #discount future cash back to NPV
        logging.info('fcf_to_bs() method complete')
        
    def fcf_to_buyback(self,price,dp = 'proportional'):
        '''
        Use cash balance to buyback shares and reduce sharecounts
        prerequisite: Must first have fcf defined
        price = share price; could be todays shareprice or anything else
        dp = 'constant' or 'proportional', 'constant' = maintain constant share price, 'proportional' = constant EV/EBITDA to todays value
        '''
        self.fin['price'] = price
        #limit buybacks to when fcf>0
        if dp == 'constant':
            for i in range(self.year):
                if i==0:
                    self.fin['buybacks'].iloc[i+1] = self.fin['fcfe'].iloc[i+1] + self.fin['cash'].iloc[0] - self.fin['dividend_policy'].iloc[i+1]
                    self.fin['shares'].iloc[i+1] = self.fin['shares'].iloc[i] - self.fin['buybacks'].iloc[i+1]/self.fin['price'].iloc[i]
                else:
                    self.fin['buybacks'].iloc[i+1] = self.fin['fcfe'].iloc[i+1] - self.fin['dividend_policy'].iloc[i+1]
                    self.fin['shares'].iloc[i+1] = self.fin['shares'].iloc[i] - self.fin['buybacks'].iloc[i+1]/self.fin['price'].iloc[i]
        elif dp == 'proportional':
            EV = price*self.shares+self.fin['debt'].iloc[0]-self.fin['cash'].iloc[0]
            multiple=EV/self.fin['ebitda'].iloc[1] #calculate the forward multiple
            for i in range(self.year-1):
                if i==0:
                    self.fin['buybacks'].iloc[i+1] = self.fin['fcfe'].iloc[i+1] + self.fin['cash'].iloc[0] - self.fin['dividend_policy'].iloc[i+1]
                    self.fin['shares'].iloc[i+1] = self.fin['shares'].iloc[i] - self.fin['buybacks'].iloc[i+1]/self.fin['price'].iloc[i]
                else:
                    self.fin['buybacks'].iloc[i+1] = self.fin['fcfe'].iloc[i+1] - self.fin['dividend_policy'].iloc[i+1]
                    self.fin['shares'].iloc[i+1] = self.fin['shares'].iloc[i] - self.fin['buybacks'].iloc[i+1]/self.fin['price'].iloc[i]
                #calculate the new price
                self.fin['price'].iloc[i+1] = (multiple*self.fin['ebitda'].iloc[i+2]-self.fin['debt'].iloc[i+1])/self.fin['shares'].iloc[i+1] #no need to include cash, since cash is being used fully for buybacks or dividend
            
            self.fin['shares'].iloc[-1] = self.fin['shares'].iloc[-2]
            self.fin['price'].iloc[-1] = self.fin['price'].iloc[-2]
        
        self.fin['dividend'] = (self.fin['fcfe']-self.fin['buybacks'])/self.fin['shares']
        self.fin['dividend'].iloc[0] = self.dividend[0]
        self.fin['dividend'].iloc[1] = (self.fin['fcfe'].iloc[1]+self.cash0-self.fin['buybacks'].iloc[1])/self.fin['shares'].iloc[1]        
        self.cash0 = 0 #all used for buybacks
        self.buybacks = True
        logging.info('fcf_to_buybacks() method complete')
        
    def fcf_to_acquire(self, adjust_cash, year_a = 1, ebitda_frac = 0.1, multiple = 10, leverage = 3, gnext = 0.1, cap_frac = 0.2):
        '''
        ebitda_frac, EBITDA of the target, relative to the organic ebitda
        gnext, next years growth
        multiple, EV/EBITDA multiple of the acquisition
        leverage, Debt/EBITDA leverage target
        adjust_cash, adjust cash balance in year 0
        '''
        
        if self.data_for_ebitda == False: 
            logging.error('financial dataset cannot be used to acquire')
        if self.fin['fcf'].empty:
            logging.error('first calculate fcf')
            
        g = [ebitda_frac-1, gnext]
        for i in range (year_a):
            g.insert(0,0)
        dEbitda = self.forecast_ebitda(self.fin['ebitda'].iloc[year_a],g)
        for i in range (year_a+1):
            dEbitda[i] = 0

        dCapex = cap_frac*np.array(dEbitda)
        dDebt = [leverage*dEbitda[year_a+1] if x >= year_a else 0 for x in range(self.year+1)]  
        self.fin['debt'] = self.fin['debt']+dDebt
        self.fin['MnA'].iloc[year_a] = self.fin['MnA'].iloc[year_a]+multiple*dEbitda[year_a+1]
        self.fin['capex'] = self.fin['capex']+dCapex
        self.fin['ebitda'] = self.fin['ebitda']+dEbitda
        self.fin['da'].iloc[year_a+1:] = np.nan
        
        if year_a == 0 and adjust_cash == True:
            self.fin['cash'].iloc[0] = self.fin['cash'].iloc[0] - (multiple-leverage)*dEbitda[1]
            self.cash0 = self.fin['cash'].iloc[0]
        
        if self.fin['cash'].iloc[year_a] < 0: 
            logging.error('cash<0, insufficient cash for the aquisition; lower the EBITDA or increase the leverage')
        
        self.fcf_from_ebitda()
        logging.info('fcf_to_acquire() method complete')
        
        return dEbitda
        
    def noa_to_dispose(self, dnoa, tax = 0, year_dis = 1, ):
        self.fin['MnA'].iloc[year_dis] = self.fin['MnA'].iloc[year_dis] - dnoa*(1-tax)
        self.fin['noa'] = self.fin['noa'] - dnoa
        self.fcf_from_ebitda()
        logging.info('dispose_from_noa() method complete')
        
    def value(self):
        '''return firm and equity values'''
        
        if self.data_for_ebitda == True:
            #really complicated way to calculate the terminal FCFE for situtions... 
            #...where there is terminal growth and you have changes in debt the final year before terminal.
            #If no change in debt and no growth the fcfet = fcfe = fcf in the terminal year
            self.fcfet = (self.fin['fcf'].iloc[-1]-self.fin['dDebt'].iloc[-1]*self.rd+self.fin['debt'].iloc[-1]*self.gt)*(1+self.gt)
            
            self.fin['equity'] = self.__pv(cfs = self.fin.fcfe, cft = self.fcfet, g = self.gt, r = self.re)            
            self.__wacc()
            self.fin['firm'] = self.__pv(cfs = self.fin.fcff, g = self.gt, r = self.fin.wacc)
            self.fin['DDM'] = self.__pv(cfs = self.fin.dividend, cft = self.fcfet/self.fin['shares'].iloc[-1], g = self.gt, r = self.re)
            
            #adjustments for cash and non-operating assets
            self.fin['value_per_share'] = (self.fin['equity']+self.fin['noa'])/self.shares
            self.fin['value_per_share'].iloc[0] = self.fin['value_per_share'].iloc[0] + self.fin['cash'].iloc[0]/self.shares
            self.fin['value_per_share_DDM'] = self.fin['DDM']+self.fin['noa']/self.fin['shares']
            self.fin['value_per_share_DDM'].iloc[0] = self.fin['value_per_share_DDM'].iloc[0]+self.cash0/self.fin['shares'].iloc[0]
            self.fin['value_per_share_DDM'].iloc[-1] = self.fin['value_per_share_DDM'].iloc[-1]+self.fin['cashBS'].iloc[-1]/self.fin['shares'].iloc[-1]

        else:
            self.fin['equity'] = self.__pv(cfs = self.fin.fcfe, g = self.gt, r = self.re)
            self.fin['firm'] = None 
        
        if self.buybacks == True:
            #calculate value per share in buyback scenario, i.e. all cash used to purchase shares until terminal year
            self.vpsbb = ((self.fin['equity'].iloc[-1]+self.fin['noa'].iloc[-1])/self.fin['shares'].iloc[-1])/(1+self.re)**self.year
        else:
            self.vpsbb = 0
            
        logging.info('value() method complete')
        return self.fin['equity'], self.fin['firm']
        
    def display_fin(self):
        
        table = self.fin[['ebitda','da','interest','income_pretax','nol','income_taxable','tax_cash','tax','capex','MnA','dDebt','dwc','fcf','fcfe','fcff','buybacks','dividend','cash','cashBS','noa','equity','debt','EV','wacc','firm','shares','price','value_per_share','value_per_share_DDM']].T.style.format("{:.1f}")
        
        wb = load_workbook(filename = os.path.join(os.path.dirname(__file__), '..\\')+'company_template.xlsx')
        ws = wb['raw data']
        
        for r in dataframe_to_rows(self.fin.T, index=True, header=True):
            ws.append(r)
        ws['A1'] = 'date'    
        
        ws = wb['report']
        ws['B2'] = self.ticker
        ws['B3'] = self.now
        ws['B4'] = 'David May'
        
        ws['B6'] = self.rd
        ws['B7'] = self.re
        ws['B8'] = self.gt
        ws['B9'] = self.t
        
        ws['B11'] = self.fin['cash'].iloc[-1]
        ws['B12'] = self.fcfet
        ws['B13'] = self.fin['equity'].iloc[-1]
        ws['B15'] = self.shares
        ws['B17'] = self.fin['value_per_share_DDM'].iloc[0]
        
        ws2 = wb['report']
        
        
        wb.save(self.ticker+'.xlsx')
        return table
